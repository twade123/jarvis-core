"""
Handler for spreadsheet operations in Excel and Numbers applications.

Capabilities:
    - Workbook management (create, open, save, close)
    - Worksheet operations (create, delete, rename)
    - Cell manipulation (read, write, select)
    - Formula application and calculations
    - Chart creation and management
    - Mathematical operations
    - Statistical functions
    - Trigonometric calculations
    - Data analysis
    - Spreadsheet formatting

Patterns:
    - "open workbook {path}"
    - "create new worksheet"
    - "write to cell {cell_ref}"
    - "apply formula to {cell_ref}"
    - "create chart from {range}"
    - "calculate {operation}"
    - "save workbook as {path}"
    - "rename worksheet to {name}"

Intents:
    - spreadsheet_workbook_manage
    - spreadsheet_worksheet_ops
    - spreadsheet_cell_ops
    - spreadsheet_formula
    - spreadsheet_chart
    - spreadsheet_calculate
    - spreadsheet_format
    - spreadsheet_analyze

Parameters:
    - command: string (spreadsheet operation)
    - spreadsheet_app: string ('Excel' or 'Numbers')
    - workbook_path: string (file path)
    - cell_ref: string (e.g., 'A1')
    - formula: string
    - operation: string (mathematical operation)
"""

import logging
import subprocess
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference

# Import agent-related components for specialized agent integration
try:
    from Jarvis_Agent_SDK.jarvis_orchestrator import analyze_handler_capabilities
    from Handler.handler_agent_builder import AgentBuilder, AgentType, AgentSpecialization, AgentCapability, AgentTool
except ImportError:
    # Allow the handler to function even if agent components can't be imported
    print("Warning: Agent components not available - specialized agent features disabled")

def osascript(script):
    """
    Execute AppleScript using osascript.
    """
    try:
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        return result.stdout.strip()
    except Exception as e:
        logging.error(f"Error executing AppleScript: {e}")
        return f"Error: {e}"

def handle_spreadsheet_intent(command, spreadsheet_app="Excel", workbook_path="example.xlsx"):
    logging.debug(f"Handling {spreadsheet_app.upper()} intent with command: {command}")

    if spreadsheet_app.lower() == "numbers":
        if command == "open_workbook":
            script = f'tell application "Numbers" to open POSIX file "{workbook_path}"'
            return osascript(script)
        
        elif command == "create_workbook":
            script = f'tell application "Numbers" to make new document at POSIX file "{workbook_path}"'
            return osascript(script)
        
        elif command == "save_workbook":
            script = f'tell application "Numbers" to save front document as POSIX file "{workbook_path}"'
            return osascript(script)
        
        elif command == "close_workbook":
            script = 'tell application "Numbers" to close front document saving yes'
            return osascript(script)
        
        elif command == "create_worksheet":
            script = 'tell application "Numbers" to make new sheet at end of sheets of front document'
            return osascript(script)
        
        elif command == "delete_worksheet":
            script = 'tell application "Numbers" to delete sheet 1 of front document'
            return osascript(script)
        
        elif command == "rename_worksheet":
            script = 'tell application "Numbers" to set name of sheet 1 of front document to "RenamedSheet"'
            return osascript(script)
        
        elif command == "select_cell":
            script = 'tell application "Numbers" to set theSelection to cell "A1" of table 1 of sheet 1 of front document'
            return osascript(script)
        
        elif command == "write_to_cell":
            script = 'tell application "Numbers" to set value of cell "A1" of table 1 of sheet 1 of front document to "Hello"'
            return osascript(script)
        
        elif command == "read_from_cell":
            script = 'tell application "Numbers" to get value of cell "A1" of table 1 of sheet 1 of front document'
            return osascript(script)
        
        elif command == "apply_formula":
            script = 'tell application "Numbers" to set formula of cell "A2" of table 1 of sheet 1 of front document to "=SUM(1, 1)"'
            return osascript(script)
        
        elif command == "create_chart":
            script = 'tell application "Numbers" to make new chart at end of charts of sheet 1 of front document'
            return osascript(script)

        elif command.startswith("calculate_"):
            operation = command.split("_")[1]
            formulas = {
                "average": "=AVERAGE(A1:A2)",
                "sum": "=SUM(A1:A3)",
                "min": "=MIN(A1:A4)",
                "max": "=MAX(A1:A5)",
                "median": "=MEDIAN(A1:A6)",
                "standard_deviation": "=STDEV(A1:A7)",
                "sin": "=SIN(A1)",
                "cos": "=COS(A1)",
                "tan": "=TAN(A1)",
                "sqrt": "=SQRT(A1)",
                "log": "=LOG(A1, 10)",
            }
            formula = formulas.get(operation)
            if formula:
                script = f'tell application "Numbers" to set formula of cell "A1" of table 1 of sheet 1 of front document to "{formula}"'
                return osascript(script)
            else:
                return f"Unknown operation '{operation}' for Numbers."

        else:
            return f"Unknown command '{command}' in Numbers context"

    elif spreadsheet_app.lower() == "excel":
        if command == "open_workbook":
            wb = load_workbook(workbook_path)
            return "Opened Excel workbook."
        
        elif command == "create_workbook":
            wb = Workbook()
            wb.save(workbook_path)
            return "Created a new Excel workbook."
        
        elif command == "save_workbook":
            wb = load_workbook(workbook_path)
            wb.save(workbook_path)
            return "Saved Excel workbook."
        
        elif command == "close_workbook":
            return "Closed Excel workbook."
        
        elif command == "create_worksheet":
            wb = load_workbook(workbook_path)
            wb.create_sheet("NewSheet")
            wb.save(workbook_path)
            return "Created a new worksheet."
        
        elif command == "delete_worksheet":
            wb = load_workbook(workbook_path)
            wb.remove(wb["NewSheet"])
            wb.save(workbook_path)
            return "Deleted a worksheet."
        
        elif command == "rename_worksheet":
            wb = load_workbook(workbook_path)
            ws = wb["NewSheet"]
            ws.title = "RenamedSheet"
            wb.save(workbook_path)
            return "Renamed a worksheet."
        
        elif command == "select_cell":
            wb = load_workbook(workbook_path)
            ws = wb.active
            cell = ws["A1"]
            return f"Selected cell {cell.coordinate}."
        
        elif command == "write_to_cell":
            wb = load_workbook(workbook_path)
            ws = wb.active
            ws["A1"] = "Hello"
            wb.save(workbook_path)
            return "Wrote to a cell."
        
        elif command == "read_from_cell":
            wb = load_workbook(workbook_path)
            ws = wb.active
            value = ws["A1"].value
            return f"Read from a cell: {value}"
        
        elif command == "apply_formula":
            wb = load_workbook(workbook_path)
            ws = wb.active
            ws["A2"] = "=SUM(1, 1)"
            wb.save(workbook_path)
            return "Applied formula to a cell."
        
        elif command == "create_chart":
            wb = load_workbook(workbook_path)
            ws = wb.active
            chart = BarChart()
            data = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=5)
            chart.add_data(data)
            ws.add_chart(chart, "E5")
            wb.save(workbook_path)
            return "Created a chart."

        elif command.startswith("calculate_"):
            operation = command.split("_")[1]
            formulas = {
                "average": "=AVERAGE(A1:A2)",
                "sum": "=SUM(A1:A3)",
                "min": "=MIN(A1:A4)",
                "max": "=MAX(A1:A5)",
                "median": "=MEDIAN(A1:A6)",
                "standard_deviation": "=STDEV(A1:A7)",
                "sin": "=SIN(A1)",
                "cos": "=COS(A1)",
                "tan": "=TAN(A1)",
                "sqrt": "=SQRT(A1)",
                "log": "=LOG(A1, 10)",
            }
            formula = formulas.get(operation)
            if formula:
                wb = load_workbook(workbook_path)
                ws = wb.active
                ws[f"A{len(ws)+1}"] = formula
                wb.save(workbook_path)
                return f"Calculated {operation.replace('_', ' ')}."
            else:
                return f"Unknown operation '{operation}' for Excel."

        else:
            return f"Unknown command '{command}' in Excel context"

    else:
        return f"Unknown spreadsheet application '{spreadsheet_app}'"